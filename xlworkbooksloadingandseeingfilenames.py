""""
from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "durga"
workbook.save(filename="hello_world.xlsx")"""
#the below snippet
from openpyxl import load_workbook
workbook = load_workbook(filename="hello_world.xlsx")
k=workbook.sheetnames
print(k)
s=workbook.active.title
l=k[0]
#print(l)
sheet = workbook[('Durga')]
sheet1 = workbook[('Sheet')]
#the below line compares the A1 values in two sheets of the workbook
print(sheet["A1"].value==sheet["A1"].value)
l=[value for value in sheet.iter_rows(values_only=True)]
f=[value for value in sheet1.iter_rows(values_only=True)]
print(l==f)
